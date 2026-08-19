# -*- coding: utf-8 -*-
"""Panen laporan keuangan RESMI IDX (XLSX ber-XBRL) -> data-idx/json/keuangan_idx/<TICKER>.json

Tugas #156. Kenapa: sumber lama (`fetch_keuangan.py`, yfinance) berlubang --
BBCA (bank terbesar di bursa) periode 2026-06-30 punya cogs/gross_profit/
operating_income/eps semuanya null, dan 313 dari 959 emiten tak punya berkas
sama sekali. XBRL IDX bertag tetap per baris dan mencakup 778 emiten TW2 2026.

SKEMA KELUARAN identik dengan `keuangan/<TICKER>.json` supaya bisa dipakai
tanpa mengubah pembacanya:

    {"ticker","currency","diperbarui","kuartal":{tanggal:{15 ruas}},"tahunan":{...}}

plus satu ruas tambahan `"sumber": "idx-xbrl"` supaya asalnya bisa dibedakan
dari berkas yfinance. Berkas ini TIDAK menimpa `keuangan/` -- direktori baru,
keputusan menggabungkan itu ada di Johan.

MATA UANG PER PERIODE (19 Agu 2026) -- tiga ruas tambahan
---------------------------------------------------------
`currency` tingkat berkas TIDAK LAGI jadi sumber kebenaran; ia tinggal
ringkasan (mata uang terbanyak). Mata uang dideklarasikan PER LAPORAN dan
penerbit boleh berganti di tengah tahun buku -- CDIA 2025/TW2 USD, 2025/TW3
IDR (sekaligus ganti skala jadi Jutaan), 2026/TW1 USD lagi -- sehingga
`Q4 = audit - TW3` menghasilkan pendapatan -64,5 TRILIUN tanpa satu pun galat.

    "mata_uang_laporan": {tanggal: "USD"}  # terbaca dari XLSX-nya (kebenaran)
    "mata_uang":         {tanggal: "USD"}  # lengkap; yang tak terbaca DITAKSIR
    "kurs_laporan":      {tanggal: 16248}  # ruas "Conversion rate", apa adanya

Pembaca yang MENGURANGKAN dua periode wajib memakai `mata_uang[<tanggal>]` dan
menolak (null) kalau beda -- sudah diterapkan di `turunkan_kuartal_diskret.py`
dan `bacaKuartalIdx` (app/src/lib/dasbor/fundamentalGabungan.ts).

`kurs_laporan` DISIMPAN tapi TIDAK dipakai mengonversi: itu kurs tanggal
pelaporan, sah untuk pos neraca saja, sedangkan pos arus (pendapatan, laba,
arus kas) butuh kurs rata-rata periode yang tak ada di sumber. Terisi di 222
dari 949 emiten; CDIA -- kasus yang memicu semua ini -- justru kosong.

CARA KERJA
----------
1. `GetFinancialReport` (docs/sumber-fundamental-idx.md #6.1) sekali panggil
   dengan `kodeEmiten` kosong mengembalikan seluruh emiten untuk satu
   (year, periode) -- jauh lebih murah daripada memanggil per emiten.
2. Unduh lampiran `FinancialStatement-*.xlsx`, parse LANGSUNG DARI MEMORI
   (`io.BytesIO`, tanpa pernah ditulis ke disk) dengan openpyxl, lalu buang.
   Tidak ada cache yang perlu di-gitignore -- byte-nya tak pernah menyentuh
   filesystem.
3. Satu XLSX berisi puluhan sheet, kodenya BEDA per industri (mis. 1xxx=Umum,
   3xxx=Infrastruktur, 4xxx=Keuangan&Syariah) -- BUKAN nomor tetap, jadi
   skrip ini mengklasifikasikan sheet lewat JUDULnya (baris pertama kolom A):
   "statement of financial position" = neraca, "statement of profit or
   loss" = laba rugi, "statement of cash flow" = arus kas. Sheet catatan
   ("notes to the financial statements ...") sengaja DILEWATI supaya
   breakdown di catatan tak terhitung ganda dengan total di laporan utama.
4. Tiap baris laporan: kolom A = label Indonesia, B = nilai periode
   berjalan, C = nilai pembanding, D = label Inggris. Dicocokkan lewat
   daftar label Inggris berprioritas (lihat *_LABELS) karena nama baris
   beda antar template industri -- bank misalnya TAK PUNYA baris "Revenue"
   sama sekali (pendapatan bunga/premi/komisi terpisah, tanpa satu baris
   total yang setara). Itu ketiadaan konsep, bukan bug -- makanya null.
5. Skala & mata uang dibaca dari sheet `1000000` ruas "Level of rounding
   used in financial statements" dan "Description of presentation
   currency" -- BUKAN diasumsikan. Diuji: BBCA/TLKM/ASII/ACST semuanya
   "Jutaan / In Million" + "Rupiah / IDR", tapi emiten tambang/lain bisa
   melapor dalam USD atau satuan penuh, jadi dibaca per-berkas.

BATASAN YANG DISADARI (baca sebelum menambah periode)
-------------------------------------------------------
- **Nilai laba-rugi & arus kas di laporan interim IDX KUMULATIF** sejak awal
  tahun buku (mis. TW2 = Jan-Jun), BUKAN kuartal diskret seperti kebiasaan
  yfinance (yang mengurangi TW2 kumulatif dengan TW1 kumulatif). Neraca tidak
  terpengaruh -- itu titik waktu (instant), bukan kumulatif. Menghitung
  diskret perlu TW1 dan TW2 sama-sama dipanen lalu diselisihkan -- belum
  dikerjakan di sini.
- "tahunan" akan KOSONG kalau hanya `--periode tw2` yang pernah dijalankan --
  laporan interim tak berisi angka tahun penuh teraudit. Isinya baru terisi
  kalau skrip ini dijalankan lagi dengan `--periode audit` (lihat "MENAMBAH
  PERIODE" di bawah).
- revenue/cogs/gross_profit/operating_income sering NULL untuk bank &
  lembaga keuangan -- taksonomi XBRL "Financial and Sharia Industry" memang
  tak punya baris "Revenue" tunggal. Diverifikasi di BBCA.
- total_debt dijumlahkan dari baris pinjaman/obligasi/efek yang diterbitkan
  di NERACA UTAMA saja (bukan sheet catatan), supaya tak dobel hitung.
  Kalau strukturnya jauh beda dari yang diuji (BBCA/TLKM/ASII/ACST), bisa
  meleset atau kosong.
- free_cf = arus kas operasi - belanja modal ("acquisition of property and
  equipment/plant"). Kalau baris belanja modal tak ketemu, free_cf ikut
  null (bukan ditebak).

TIDAK BOLEH DIJALANKAN DARI GITHUB ACTIONS / RUNNER DATACENTER -- endpoint
IDX menjawab 403 dari sana, terverifikasi 16 Agu 2026
(docs/sumber-fundamental-idx.md). Hanya jalan dari IP rumahan
(JALANKAN_OTOMATIS.bat / manual).

PAKAI
-----
  python scripts/panen_keuangan_idx.py --tickers BBCA,TLKM,ASII
  python scripts/panen_keuangan_idx.py --semua                  # 778 emiten TW2 2026
  python scripts/panen_keuangan_idx.py --semua --paksa          # ulang walau sudah ada
  python scripts/panen_keuangan_idx.py --semua --periode tw1    # tambah TW1
  python scripts/panen_keuangan_idx.py --swauji                 # uji penaksir mata uang
  python scripts/panen_keuangan_idx.py --dari-arsip --tahun 2025 --periode tw3

`--dari-arsip` memeras ULANG dari `_arsip-mentah/` -- NOL permintaan jaringan.
Ia hanya MENAMBAH: mata uang per periode selalu diperbarui, tapi nilai periode
ditulis hanya kalau periode itu belum ada. Dua alasan: (a) periode yang
mentahnya tak pernah terarsip (2020/2021/2025-audit, 2026-TW2, ~2.300 catatan)
tak boleh ikut terhapus, dan (b) nilai yang ada sudah lewat
`perbaiki_skala_keuangan.py`. Karena itu `--dari-arsip` aman dijalankan atas
seluruh arsip (`--semua-arsip`), berbeda dari `--paksa` di jalur jaringan.

Emiten yang berkasnya SUDAH punya data untuk (tanggal, bucket) yang diminta
dilewati otomatis -- resumable kalau 778 emiten putus di tengah jalan, dan
sekaligus jadi cara menambah periode lain (lihat di bawah).

MENAMBAH PERIODE LAIN (pekerjaan berikutnya)
---------------------------------------------
Jalankan lagi dengan `--periode`/`--tahun` berbeda. Skrip MENGGABUNG ke
berkas yang sudah ada (union per tanggal), bukan menimpa seluruh berkas --
jadi TW1+TW2+TW3+audit bisa ditumpuk dengan menjalankan berkali-kali.
`--periode audit` mengisi bucket "tahunan"; `tw1`/`tw2`/`tw3` mengisi
"kuartal". `--paksa` memaksa refetch walau tanggal itu sudah ada (mis. kalau
emiten merevisi laporan).
"""
from __future__ import annotations

import argparse
import io
import json
import math
import random
import sys
import time
import urllib.parse
from collections import Counter
from datetime import datetime, timedelta, timezone
from pathlib import Path

import requests  # noqa: F401 -- dipakai untuk tipe galat di penanganan retry
from curl_cffi import requests as cffi
from openpyxl import load_workbook

# Sidik jari peramban yang ditiru. Diganti kalau IDX mulai menolak lagi;
# daftar nilai yang sah ada di dokumentasi curl_cffi (chrome110..chrome124, dst).
IMPERSONATE = "chrome124"

sys.path.insert(0, str(Path(__file__).resolve().parent))
from fetch_fundamental import DEFAULT_TICKERS  # reuse -- lihat CLAUDE.md rung 2

AKAR = Path(__file__).resolve().parent.parent
KELUARAN_DIR = AKAR / "data-idx" / "json" / "keuangan_idx"
WIB = timezone(timedelta(hours=7))

UA = ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
      "(KHTML, like Gecko) Chrome/140.0.0.0 Safari/537.36")

# Header ala peramban SUNGGUHAN, bukan tiga baris seadanya (18 Agu 2026).
#
# Sebelumnya header cuma {User-Agent, Referer: /id, Accept: application/json},
# dan seluruh panggilan mulai dijawab 403 -- termasuk tahun 2025 yang beberapa
# jam sebelumnya berhasil dipanen 882 emiten. Dugaan pertama "IP-nya diblokir"
# SALAH, dan itu terbukti dalam satu langkah: URL yang sama persis dibuka di
# Chrome menjawab 200 dengan 664 hasil untuk AUDIT 2019. Yang membedakan bukan
# alamatnya, melainkan bentuk permintaannya.
#
# Yang menyembuhkan: kumpulan header yang benar-benar dikirim peramban --
# Accept-Language, sec-ch-ua*, Sec-Fetch-*, dan Referer yang menunjuk ke
# HALAMAN LAPORAN KEUANGAN, bukan ke beranda. Sesudah ini `requests` biasa
# menjawab 200 lagi tanpa perlu pustaka apa pun.
#
# Kalau suatu saat ini pun ditolak, jalan berikutnya `curl_cffi` (sudah
# terpasang, versi 0.16.0) dengan `impersonate="chrome124"` -- ia meniru sidik
# jari TLS, satu lapis lebih dalam daripada header. Sudah diuji hari ini dan
# juga menjawab 200. Jangan menaikkan laju permintaan sebagai jalan keluar;
# yang ditolak bentuk permintaannya, bukan jumlahnya.
_HDR_PERAMBAN = {
    "User-Agent": UA,
    "Accept-Language": "id-ID,id;q=0.9,en-US;q=0.8,en;q=0.7",
    "sec-ch-ua": '"Chromium";v="140", "Not=A?Brand";v="24", "Google Chrome";v="140"',
    "sec-ch-ua-mobile": "?0",
    "sec-ch-ua-platform": '"Windows"',
    "Sec-Fetch-Dest": "empty",
    "Sec-Fetch-Mode": "cors",
    "Sec-Fetch-Site": "same-origin",
    "Referer": "https://www.idx.co.id/id/perusahaan-tercatat/laporan-keuangan-dan-tahunan/",
}
HEADER = {**_HDR_PERAMBAN, "Accept": "application/json, text/plain, */*"}
# Unduhan xlsx GAGAL 406 kalau dikirim header "Accept: application/json" yang
# sama dengan panggilan API JSON -- perlu header terpisah tanpa itu.
HEADER_FILE = {**_HDR_PERAMBAN}
PEMANASAN = "https://www.idx.co.id/id"
LIST_URL = "https://www.idx.co.id/primary/ListedCompany/GetFinancialReport"
HOST = "https://www.idx.co.id"

JEDA = (1.0, 2.2)               # jeda acak antar unduhan xlsx, detik
TUNGGU_ULANG = (5, 15, 45)      # backoff kalau 403/429/gagal jaringan

PERIODE_AKHIR = {"tw1": "03-31", "tw2": "06-30", "tw3": "09-30", "audit": "12-31"}

ALL_KEYS = [
    "revenue", "cogs", "gross_profit", "operating_income", "net_income", "eps",
    "total_assets", "total_liabilities", "equity", "cash", "total_debt",
    "operating_cf", "investing_cf", "financing_cf", "free_cf",
]

# Label Inggris (XBRL StandardLabel), berprioritas -- yang pertama ketemu
# (dengan nilai bukan None) di sheet bertipe terkait yang menang.
LABARUGI_REVENUE = ["Sales and revenue", "Net revenue", "Revenue"]
LABARUGI_COGS = ["Cost of sales and revenue", "Cost of revenue"]
LABARUGI_GROSS = ["Total gross profit", "Gross profit"]
LABARUGI_OPINC = [
    "Total profit from operation", "Total profit (loss) from operation",
    "Profit (loss) from operation", "Operating profit (loss)",
]
LABARUGI_NETINC = ["Profit (loss) attributable to parent entity", "Total profit (loss)"]
LABARUGI_EPS = [
    "Basic earnings per share attributable to equity owners of the parent entity",
    "Basic earnings (loss) per share from continuing operations",
    "Earnings (loss) per share",
]
NERACA_ASET = ["Total assets"]
NERACA_LIAB = ["Total liabilities"]
NERACA_EKUITAS = ["Total equity"]
NERACA_KAS = ["Cash and cash equivalents"]
KAS_OPERASI = ["Total net cash flows received from (used in) operating activities"]
KAS_INVESTASI = ["Total net cash flows received from (used in) investing activities"]
KAS_PENDANAAN = ["Total net cash flows received from (used in) financing activities"]
KAS_AKHIR = ["Cash and cash equivalents cash flows, end of the period"]
KAS_CAPEX = [
    "Payments for acquisition of property and equipment",
    "Payments for acquisition of property, plant and equipment",
]
# Dijumlahkan (bukan diprioritaskan) -- baris-baris pinjaman/obligasi/efek
# berbunga di NERACA UTAMA. Aman dijumlahkan karena hanya SATU varian sheet
# neraca yang terisi per emiten (lihat cari_peta), jadi tak dobel hitung.
NERACA_DEBT_LABELS = [
    "Short term bank loans", "Long-term bank loans", "Current maturities of bank loans",
    "Bonds payable", "Long-term bonds payable", "Current maturities of bonds payable",
    "Other borrowings", "Long-term other borrowings", "Current maturities of other borrowings",
    "Borrowings third parties", "Borrowings related parties",
    "Others securities issued", "Bank securities issued",
    "Subordinated bonds", "Subordinated loans third parties", "Subordinated loans related parties",
    "Subordinated mudharabah sukuk",
]


class Ditolak(Exception):
    """IDX menolak berturut-turut (403/429) -- isyarat berhenti."""


def tanggal_akhir(tahun: int, periode: str) -> str:
    return f"{tahun}-{PERIODE_AKHIR[periode]}"


def klasifikasi_judul(judul: str) -> str | None:
    j = judul.lower()
    if "statement of financial position" in j:
        return "neraca"
    if "statement of profit or loss" in j:
        return "labarugi"
    if "statement of cash flow" in j:
        return "kas"
    return None


def peta_dari_tipe(wb, tipe: str, kolom: int = 1) -> dict[str, float]:
    """Gabungkan seluruh sheet bertipe `tipe` jadi satu peta label_en(lower) -> nilai
    periode berjalan (`kolom=1`, kolom B) atau periode PEMBANDING (`kolom=2`,
    kolom C -- lihat `panen_pembanding.py`).

    Hanya sheet yang JUDULnya cocok yang disentuh (sheet
    catatan/"notes" otomatis terlewat). Kalau ada beberapa varian sheet untuk
    tipe yang sama (mis. "by function" vs "by nature"), hanya satu yang
    benar-benar terisi angka -- yang lain tetap None dan tak menimpa."""
    peta: dict[str, float] = {}
    for nama in wb.sheetnames:
        ws = wb[nama]
        baris_iter = ws.iter_rows(values_only=True)
        try:
            judul = next(baris_iter)[0]
        except StopIteration:
            continue
        if not isinstance(judul, str) or klasifikasi_judul(judul) != tipe:
            continue
        for row in baris_iter:
            if len(row) < 4:
                continue
            label_en, nilai = row[3], row[kolom]
            if isinstance(label_en, str) and nilai is not None:
                kunci = label_en.strip().lower()
                if kunci not in peta:
                    peta[kunci] = nilai
    return peta


def cari(peta: dict, kandidat: list[str]):
    for lbl in kandidat:
        v = peta.get(lbl.lower())
        if v is not None:
            return v
    return None


def info_umum(wb) -> tuple[str, int, float | None]:
    """(currency, skala, kurs) dari sheet 1000000 -- dibaca, bukan diasumsikan.

    `kurs` = ruas "Conversion rate at reporting date". Diukur 19 Agu 2026 atas
    SELURUH arsip mentah: tak satu pun berkas mengisinya. Tetap dibaca supaya
    kalau suatu saat terisi, konversi lintas mata uang punya dasar dari sumber
    dan bukan angka karangan.
    """
    currency, skala, kurs = "IDR", 1_000_000, None
    if "1000000" not in wb.sheetnames:
        return currency, skala, kurs
    for row in wb["1000000"].iter_rows(values_only=True):
        if len(row) < 3 or not isinstance(row[2], str):
            continue
        label_en = row[2].lower()
        if "description of presentation currency" in label_en:
            val = str(row[1] or "")
            if "/" in val:
                currency = val.split("/")[-1].strip().upper() or "IDR"
        elif "conversion rate at reporting date" in label_en:
            if isinstance(row[1], (int, float)) and not isinstance(row[1], bool) and row[1]:
                kurs = float(row[1])
        elif "level of rounding" in label_en:
            val = str(row[1] or "").lower()
            # Diuji: BBCA/ASII/ACST = "Jutaan/Million", TLKM = "Miliaran/Billion"
            # -- keduanya nyata dipakai, jangan asumsikan cuma satu.
            if "miliar" in val or "billion" in val:
                skala = 1_000_000_000
            elif "juta" in val or "million" in val:
                skala = 1_000_000
            elif "ribu" in val or "thousand" in val:
                skala = 1_000
            else:
                skala = 1
    return currency, skala, kurs


def ekstrak(wb, kolom: int = 1) -> tuple[dict, str, float | None]:
    """`kolom=1` (kolom B) = periode berjalan; `kolom=2` (kolom C) = periode
    PEMBANDING. Skala & mata uang tetap dibaca dari sheet 1000000 kolom B --
    deklarasi itu berlaku untuk seluruh berkas, termasuk kolom pembandingnya
    (IAS 21: komparatif disajikan ulang dalam mata uang penyajian berjalan)."""
    currency, skala, kurs = info_umum(wb)

    def rp(v):
        if v is None:
            return None
        try:
            return float(v) * skala
        except (TypeError, ValueError):
            return None

    def rp_eps(v):
        # ponytail: EPS seharusnya (spesifikasi XBRL) TIDAK ikut skala moneter
        # entitas -- elemen perShareItemType selalu dalam mata uang penuh per
        # lembar, beda dari monetaryItemType. Tapi diverifikasi: sebagian
        # emiten (BBCA/BMRI/ICBP) menandainya sudah dalam rupiah penuh,
        # sebagian lain (TLKM/ASII/ACST/ALKA/AKKU/ADES) tetap ikut skala
        # jutaan -- inkonsistensi di berkas sumber, bukan satu aturan yang
        # berlaku semua. Diputuskan pakai ambang plausibilitas: pilih varian
        # (mentah vs dikali skala) yang jatuh di rentang EPS wajar saham IDX.
        # Upgrade path kalau ini kurang -- baca "decimals" dari instance.zip
        # XBRL mentah per fakta, bukan dari nilai XLSX yang sudah dirender.
        if v is None:
            return None
        try:
            mentah = float(v)
        except (TypeError, ValueError):
            return None
        skalaan = mentah * skala
        PLAUSIBEL = (0.001, 1_000_000)
        if PLAUSIBEL[0] <= abs(mentah) <= PLAUSIBEL[1]:
            return mentah
        if PLAUSIBEL[0] <= abs(skalaan) <= PLAUSIBEL[1]:
            return skalaan
        return mentah  # tak satu pun wajar -- pakai mentah, lebih jujur dari menebak

    neraca = peta_dari_tipe(wb, "neraca", kolom)
    labarugi = peta_dari_tipe(wb, "labarugi", kolom)
    kas = peta_dari_tipe(wb, "kas", kolom)

    # `is not None` meloloskan STRING, dan satu sel neraca berisi teks membuat
    # sum() melempar TypeError yang menggagalkan seluruh emiten. Terukur pada
    # AUDIT 2022: 246 emiten gagal karenanya -- persis jumlah yang selama ini
    # hilang dari periode itu. Sel berteks dilewati, bukan dijadikan nol: nol
    # berarti "utangnya nol", dan itu pernyataan yang belum tentu benar.
    total_debt_raw = sum(
        v
        for v in (neraca.get(lbl.lower()) for lbl in NERACA_DEBT_LABELS)
        if isinstance(v, (int, float)) and not isinstance(v, bool)
    )
    total_debt = rp(total_debt_raw) if total_debt_raw else None

    operating_cf = rp(cari(kas, KAS_OPERASI))
    investing_cf = rp(cari(kas, KAS_INVESTASI))
    financing_cf = rp(cari(kas, KAS_PENDANAAN))
    capex = rp(cari(kas, KAS_CAPEX))
    free_cf = (operating_cf - capex) if (operating_cf is not None and capex is not None) else None

    cash = rp(cari(neraca, NERACA_KAS))
    if cash is None:
        cash = rp(cari(kas, KAS_AKHIR))  # bank: tak ada baris neraca "cash and cash equivalents"

    data = {
        "revenue": rp(cari(labarugi, LABARUGI_REVENUE)),
        "cogs": rp(cari(labarugi, LABARUGI_COGS)),
        "gross_profit": rp(cari(labarugi, LABARUGI_GROSS)),
        "operating_income": rp(cari(labarugi, LABARUGI_OPINC)),
        "net_income": rp(cari(labarugi, LABARUGI_NETINC)),
        "eps": rp_eps(cari(labarugi, LABARUGI_EPS)),
        "total_assets": rp(cari(neraca, NERACA_ASET)),
        "total_liabilities": rp(cari(neraca, NERACA_LIAB)),
        "equity": rp(cari(neraca, NERACA_EKUITAS)),
        "cash": cash,
        "total_debt": total_debt,
        "operating_cf": operating_cf,
        "investing_cf": investing_cf,
        "financing_cf": financing_cf,
        "free_cf": free_cf,
    }
    return data, currency, kurs


def ambil_daftar(sesi: requests.Session, tahun: int, periode: str) -> dict:
    params = {
        "indexFrom": 1, "pageSize": 1000, "year": tahun, "reportType": "rdf",
        "EmitenType": "s", "periode": periode, "kodeEmiten": "",
        "SortColumn": "KodeEmiten", "SortOrder": "asc",
    }
    # 403 di endpoint ini TRANSIEN, bukan penolakan menetap. Terbukti 17 Agu
    # 2026: dalam satu proses yang sama, panen AUDIT 2020 lolos penuh (699
    # emiten), 2021 mulai tersendat (26 gagal), lalu 2022-2025 gagal SEMUA di
    # langkah pengambilan daftar ini. Diuji ulang beberapa menit kemudian dari
    # IP yang sama: tahun 2022 masih 403 tapi 2025 menjawab 200 dengan 882
    # hasil -- jadi yang terjadi pembatasan laju, bukan pemblokiran.
    #
    # Tanpa percobaan ulang, SATU 403 membatalkan seluruh tahun. Padahal
    # biayanya cuma menunggu: jeda menaik 5, 15, 30, 60 detik.
    jeda = [5, 15, 30, 60]
    galat_terakhir: Exception | None = None
    for percobaan in range(len(jeda) + 1):
        try:
            r = sesi.get(LIST_URL, params=params, headers=HEADER, timeout=60)
            r.raise_for_status()
            break
        except Exception as e:  # noqa: BLE001
            galat_terakhir = e
            if percobaan >= len(jeda):
                raise
            tunggu = jeda[percobaan]
            print(f"    daftar {periode.upper()} {tahun} ditolak ({e.__class__.__name__}) "
                  f"-- coba lagi dalam {tunggu} detik "
                  f"[{percobaan + 1}/{len(jeda)}]", flush=True)
            time.sleep(tunggu)
    else:  # pragma: no cover - dijaga `raise` di atas
        raise galat_terakhir or RuntimeError("daftar laporan tak terambil")
    hasil = {}
    for entri in (r.json().get("Results") or []):
        kode = (entri.get("KodeEmiten") or "").strip().upper()
        if kode:
            hasil[kode] = entri
    return hasil


def cari_xlsx(entri: dict) -> dict | None:
    for a in entri.get("Attachments") or []:
        nama = (a.get("File_Name") or "")
        if nama.lower().endswith(".xlsx") and nama.lower().startswith("financialstatement"):
            return a
    return None


# Arsip berkas MENTAH. Sengaja di luar repo (lihat .gitignore) supaya 200-an MB
# XLSX tak ikut tiap kloning, tapi TETAP ADA di cakram.
#
# Johan 17 Agu 2026: "jangan asal maen buang data yang sudah di panen, gini ini
# jadi masalah kan harus unduh lagi, simpan backup saja sewaktu perlu kita
# gunakan gini". Latar belakangnya nyata: versi pertama skrip ini memeras XLSX
# jadi 15 ruas lalu membuang berkasnya. Begitu muncul kebutuhan ruas lain
# (neraca saja 238 baris), satu-satunya jalan adalah mengunduh ulang seluruh
# 900-an emiten — padahal endpoint-nya sering menolak 403 dan panen penuh makan
# berjam-jam.
#
# Aturan yang berlaku sejak sekarang: yang mahal itu MENGAMBILNYA, bukan
# menyimpannya. Simpan mentahnya, parse-nya boleh diulang kapan saja.
ARSIP_MENTAH = AKAR / "_arsip-mentah" / "keuangan_idx"


def jalur_arsip(kode: str, tahun: int, periode: str) -> Path:
    return ARSIP_MENTAH / str(tahun) / periode / f"{kode}.xlsx"


def ambil_xlsx(sesi: requests.Session, file_path: str, kode: str,
               tahun: int, periode: str) -> bytes:
    """Isi XLSX — dari arsip cakram kalau ada, kalau tidak baru diunduh.

    Ini yang membuat penambahan ruas di kemudian hari tak berbiaya jaringan
    sama sekali: jalankan ulang dengan `--paksa`, seluruh berkas dibaca dari
    cakram, tak satu pun permintaan keluar.
    """
    berkas = jalur_arsip(kode, tahun, periode)
    if berkas.exists() and berkas.stat().st_size > 0:
        return berkas.read_bytes()
    konten = unduh_xlsx(sesi, file_path)
    try:
        berkas.parent.mkdir(parents=True, exist_ok=True)
        berkas.write_bytes(konten)
    except OSError as e:  # cakram penuh / izin -- panen JANGAN ikut gagal
        print(f" [arsip gagal: {e}]", end="")
    return konten


def unduh_xlsx(sesi: requests.Session, file_path: str) -> bytes:
    url = HOST + urllib.parse.quote(file_path)
    galat: Exception | None = None
    for n, jeda in enumerate((0, *TUNGGU_ULANG)):
        if jeda:
            time.sleep(jeda)
        try:
            r = sesi.get(url, headers=HEADER_FILE, timeout=60)
            if r.status_code in (403, 429):
                galat = Ditolak(f"HTTP {r.status_code}")
                continue
            r.raise_for_status()
            return r.content
        except requests.RequestException as e:
            galat = e
    raise galat or RuntimeError("gagal mengunduh tanpa alasan jelas")


def sudah_ada(kode: str, tanggal: str, bucket: str) -> bool:
    path = KELUARAN_DIR / f"{kode}.json"
    if not path.exists():
        return False
    try:
        isi = json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return False
    return tanggal in (isi.get(bucket) or {})


def _ordinal(tanggal: str) -> int:
    return datetime.strptime(tanggal, "%Y-%m-%d").toordinal()


def _aset(periode) -> float | None:
    v = (periode or {}).get("total_assets")
    if isinstance(v, (int, float)) and not isinstance(v, bool) and v:
        return abs(float(v))
    return None


def lengkapi_mata_uang(isi: dict) -> dict[str, str]:
    """Peta LENGKAP tanggal -> mata uang, satu entri untuk tiap periode.

    Kenapa ada (19 Agu 2026, CDIA): mata uang dideklarasikan PER LAPORAN, dan
    penerbit berganti di tengah tahun buku. CDIA 2025/TW2 USD, 2025/TW3 IDR,
    2026/TW1 USD lagi. Satu `currency` tingkat berkas menyembunyikan itu, dan
    `Q4 = audit - TW3` lalu menghasilkan revenue -64,5 triliun tanpa satu pun
    galat.

    `mata_uang_laporan` = yang benar-benar terbaca dari berkas XLSX-nya
    (kebenaran). Sebagian periode dipanen sebelum arsip mentah ada, jadi
    berkasnya tak bisa dibaca lagi (2020/2021/2025-audit, 2026-TW2) -- untuk itu
    mata uangnya DITAKSIR dari dua isyarat, urut:

    1. JANGKAR NERACA, tapi hanya kalau JAWABANNYA TEGAS. Ganti mata uang
       rupiah<->dolar itu lompatan ~4,2 dekade besaran; jadi jangkar dipakai
       hanya kalau ada satu mata uang yang `total_assets`-nya sepadan (dalam
       10x) SEKALIGUS semua mata uang lain jauh (di luar 100x). Tanpa syarat
       ketegasan itu, SATU periode yang cacat di sumbernya menular: SGER
       melaporkan 2023 sebagai USD dengan angka yang bukan USD maupun IDR, dan
       jangkar polos lalu ikut melabeli 2020 & 2021 sebagai USD padahal
       keduanya jelas rupiah.
    2. Kalau jangkar tak tegas: TANGGAL TERDEKAT (seri -> yang lebih baru).

    Alasan jangkar neraca dipilih lebih dulu daripada tanggal (pola yang sama
    dipakai `perbaiki_skala_keuangan.py`): neraca itu posisi pada satu tanggal,
    bergerak lambat. Tanggal saja TIDAK cukup: CDIA 2025-12-31 berjarak sama (3
    bulan) dari 2025-09-30 (IDR) dan 2026-03-31 (USD) dan yang benar USD; ANJT
    2025-12-31 juga seri, tapi yang benar justru tetangga yang lebih tua.
    Jangkar menjawab keduanya benar; tanggal menjawab salah satu salah, pilihan
    mana pun.

    Kalau seluruh periode terbaca bermata uang sama, tak ada yang perlu
    ditaksir -- itu 936 dari 949 berkas.
    """
    laporan: dict[str, str] = isi.get("mata_uang_laporan") or {}
    if not laporan:
        return {}
    semua: dict[str, dict] = {}
    for b in ("kuartal", "tahunan"):
        semua.update(isi.get(b) or {})

    hasil = {t: c for t, c in laporan.items() if t in semua}
    if len(set(laporan.values())) == 1:
        satu = next(iter(laporan.values()))
        return {t: hasil.get(t, satu) for t in semua}

    # Sepadan kalau dalam 10x (1 dekade); jauh kalau di luar 100x (2 dekade).
    # Ganti mata uang IDR<->USD sendiri ~4,2 dekade, jadi celah 1..2 itu lebar.
    DEKAT, JAUH = 1.0, 2.0

    def terdekat_tanggal(t: str) -> str:
        return min(laporan, key=lambda k: (abs(_ordinal(k) - _ordinal(t)), -_ordinal(k)))

    for t in semua:
        if t in hasil:
            continue
        a_t = _aset(semua[t])
        jarak: dict[str, float] = {}
        if a_t:
            for k, c in laporan.items():
                a_k = _aset(semua.get(k))
                if a_k:
                    d = abs(math.log10(a_t / a_k))
                    jarak[c] = min(jarak.get(c, d), d)
        menang = min(jarak, key=jarak.get) if jarak else None
        tegas = (
            menang is not None
            and jarak[menang] < DEKAT
            and all(d > JAUH for c, d in jarak.items() if c != menang)
        )
        hasil[t] = menang if tegas else laporan[terdekat_tanggal(t)]
    return hasil


def swauji_mata_uang() -> None:
    """Swauji `lengkapi_mata_uang` -- bentuk nyata dari empat emiten yang
    masing-masing mematahkan satu versi sebelumnya dari penaksirnya."""
    def bikin(aset: dict[str, float], laporan: dict[str, str]) -> dict:
        return {"kuartal": {}, "tahunan": {t: {"total_assets": v} for t, v in aset.items()},
                "mata_uang_laporan": laporan}

    # CDIA: seri tanggal (3 bulan ke belakang IDR, 3 bulan ke depan USD).
    # Jangkar tegas -> USD, dan itu yang benar.
    cdia = bikin(
        {"2025-06-30": 1.39e9, "2025-09-30": 1.68e14, "2025-12-31": 1.74e9, "2026-03-31": 1.90e9},
        {"2025-06-30": "USD", "2025-09-30": "IDR", "2026-03-31": "USD"})
    assert lengkapi_mata_uang(cdia)["2025-12-31"] == "USD"

    # ANJT: seri tanggal juga, tapi jawabannya tetangga yang LEBIH TUA.
    anjt = bikin(
        {"2025-09-30": 5.95e8, "2025-12-31": 3.17e8, "2026-03-31": 6.87e12, "2026-06-30": 7.05e12},
        {"2025-09-30": "USD", "2026-03-31": "IDR"})
    hasil = lengkapi_mata_uang(anjt)
    assert hasil["2025-12-31"] == "USD", hasil
    assert hasil["2026-06-30"] == "IDR", hasil

    # SGER: satu periode CACAT di sumber (2023 berlabel USD dengan angka yang
    # bukan USD maupun IDR). Jangkar tak tegas -> jatuh ke tanggal terdekat.
    sger = bikin(
        {"2020-12-31": 6.86e11, "2021-12-31": 1.24e12, "2022-12-31": 3.37e12,
         "2023-12-31": 2.13e11, "2024-12-31": 4.64e12},
        {"2022-12-31": "IDR", "2023-12-31": "USD", "2024-12-31": "IDR"})
    hasil = lengkapi_mata_uang(sger)
    assert hasil["2020-12-31"] == "IDR", hasil
    assert hasil["2021-12-31"] == "IDR", hasil

    # Seluruh periode terbaca satu mata uang -> tak ada yang ditaksir.
    seragam = bikin({"2024-12-31": 1e12, "2025-12-31": 1.1e12}, {"2024-12-31": "IDR"})
    assert lengkapi_mata_uang(seragam) == {"2024-12-31": "IDR", "2025-12-31": "IDR"}

    # Belum pernah ada laporan terbaca -> peta kosong, BUKAN tebakan "IDR".
    assert lengkapi_mata_uang(bikin({"2024-12-31": 1e12}, {})) == {}
    print("lengkapi_mata_uang: swauji lolos")


def _dominan(peta: dict[str, str], cadangan: str) -> str:
    if not peta:
        return cadangan
    return Counter(peta.values()).most_common(1)[0][0]


def simpan(kode: str, tanggal: str, bucket: str, data_periode: dict | None,
           currency: str, kurs: float | None = None) -> None:
    """Tulis satu periode. `data_periode=None` -> hanya perbarui mata uang
    (dipakai `--dari-arsip`: nilainya sudah ada dan sudah lewat penambalan
    skala, jadi jangan ditimpa)."""
    KELUARAN_DIR.mkdir(parents=True, exist_ok=True)
    path = KELUARAN_DIR / f"{kode}.json"
    isi = {}
    if path.exists():
        try:
            isi = json.loads(path.read_text(encoding="utf-8"))
        except Exception:
            isi = {}
    isi["ticker"] = kode
    isi["sumber"] = "idx-xbrl"
    isi.setdefault("kuartal", {})
    isi.setdefault("tahunan", {})
    if data_periode is not None:
        isi[bucket][tanggal] = data_periode
    if currency and tanggal in (isi.get(bucket) or {}):
        isi.setdefault("mata_uang_laporan", {})[tanggal] = currency
        # Kurs SUMBER apa adanya (rupiah per satuan mata uang pelaporan, pada
        # TANGGAL PELAPORAN). Disimpan, TIDAK dipakai mengonversi: kurs tanggal
        # penutupan sah untuk pos neraca saja; pos ARUS (pendapatan, laba, arus
        # kas) semestinya pakai kurs rata-rata periode, yang tak ada di sumber.
        # Mengonversi dengan satu-satunya kurs yang ada akan menghasilkan angka
        # yang terlihat presisi dan salah -- persis cacat yang sedang dibetulkan.
        if kurs:
            isi.setdefault("kurs_laporan", {})[tanggal] = kurs
    isi["mata_uang"] = lengkapi_mata_uang(isi)
    # `currency` tingkat berkas tinggal RINGKASAN (mata uang terbanyak), bukan
    # lagi satu-satunya sumber kebenaran -- pembaca yang mengurangkan wajib
    # memakai `mata_uang[<tanggal>]`.
    isi["currency"] = _dominan(isi["mata_uang"], currency or isi.get("currency") or "IDR")
    isi["diperbarui"] = datetime.now(WIB).strftime("%Y-%m-%d %H:%M")
    path.write_text(json.dumps(isi, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")


def segarkan_mata_uang(kode: str) -> bool:
    """Hitung ulang `mata_uang` dari `mata_uang_laporan` yang sudah tersimpan.
    Dipakai di akhir `--dari-arsip` supaya berkas yang periodenya baru sebagian
    diperas ulang tetap punya peta lengkap."""
    path = KELUARAN_DIR / f"{kode}.json"
    try:
        isi = json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return False
    baru = lengkapi_mata_uang(isi)
    if baru == (isi.get("mata_uang") or {}):
        return False
    isi["mata_uang"] = baru
    isi["currency"] = _dominan(baru, isi.get("currency") or "IDR")
    path.write_text(json.dumps(isi, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
    return True


def jalankan_arsip(tickers: set[str] | None, tahun: int | None, periode: str | None) -> int:
    """Peras ULANG dari `_arsip-mentah/` -- NOL permintaan jaringan.

    Hanya menambah: mata uang per periode selalu diperbarui, tapi NILAI periode
    ditulis hanya kalau periode itu belum ada. Dua alasan, keduanya sudah
    dibayar: (a) menjalankan ulang atas SELURUH arsip tak boleh sampai
    menghapus periode yang mentahnya tak pernah terarsip (2020/2021/2025-audit,
    2026-TW2 -- ~2.300 catatan), dan (b) nilai yang ada sudah lewat
    `perbaiki_skala_keuangan.py`, menimpanya akan mengembalikan cacat skalanya.
    """
    berkas_arsip = sorted(ARSIP_MENTAH.glob("*/*/*.xlsx"))
    diproses = ditambah = gagal = 0
    kurs_ada: set[str] = set()
    for f in berkas_arsip:
        th, per, kode = int(f.parent.parent.name), f.parent.name, f.stem.upper()
        if tahun and th != tahun:
            continue
        if periode and per != periode:
            continue
        if tickers and kode not in tickers:
            continue
        if per not in PERIODE_AKHIR:
            continue
        bucket = "tahunan" if per == "audit" else "kuartal"
        tanggal = tanggal_akhir(th, per)
        try:
            wb = load_workbook(io.BytesIO(f.read_bytes()), data_only=True, read_only=True)
            data_periode, currency, kurs = ekstrak(wb)
            wb.close()
        except Exception as e:  # noqa: BLE001
            gagal += 1
            print(f"  gagal {kode} {th}/{per}: {e}")
            continue
        diproses += 1
        if kurs:
            kurs_ada.add(kode)
        baru = None
        if not sudah_ada(kode, tanggal, bucket):
            if sum(1 for v in data_periode.values() if v is not None) == 0:
                continue  # semua ruas null -- jangan menulis catatan kosong
            baru = data_periode
            ditambah += 1
        simpan(kode, tanggal, bucket, baru, currency, kurs)
    disegarkan = sum(segarkan_mata_uang(p.stem) for p in sorted(KELUARAN_DIR.glob("*.json")))
    print(f"\nArsip: {diproses} berkas diperas ulang, {ditambah} periode BARU ditulis, "
          f"{gagal} gagal, {disegarkan} berkas peta mata uangnya disegarkan")
    print(f"Kurs (Conversion rate) terisi di sumber: {len(kurs_ada)} emiten -- "
          "DISIMPAN sebagai `kurs_laporan`, tidak dipakai mengonversi (lihat simpan())")
    return 0


def main() -> int:
    ap = argparse.ArgumentParser(description="Panen laporan keuangan resmi IDX (XBRL)")
    ap.add_argument("--tickers", help="Daftar kode dipisah koma, mis. BBCA,TLKM")
    ap.add_argument("--semua", action="store_true", help="Seluruh emiten (DEFAULT_TICKERS)")
    ap.add_argument("--paksa", action="store_true", help="Panen ulang walau periode sudah ada")
    ap.add_argument("--periode", default="tw2", choices=["tw1", "tw2", "tw3", "audit"])
    ap.add_argument("--tahun", type=int, default=2026)
    ap.add_argument("--batas", type=int, default=None, help="Batasi jumlah tiker (uji cepat)")
    ap.add_argument("--dari-arsip", action="store_true",
                    help="Peras ulang dari _arsip-mentah/ (nol jaringan). Tanpa "
                         "--tahun/--periode: seluruh arsip yang ada.")
    ap.add_argument("--semua-arsip", action="store_true",
                    help="Bersama --dari-arsip: seluruh tahun & periode yang ada di arsip")
    ap.add_argument("--swauji", action="store_true",
                    help="Uji penaksir mata uang per periode, tak menyentuh data")
    ap.add_argument("--segarkan-mata-uang", action="store_true",
                    help="Hitung ulang peta `mata_uang` seluruh berkas dari `mata_uang_laporan`")
    args = ap.parse_args()

    if args.swauji:
        swauji_mata_uang()
        return 0

    if args.segarkan_mata_uang:
        n = sum(segarkan_mata_uang(p.stem) for p in sorted(KELUARAN_DIR.glob("*.json")))
        print(f"{n} berkas peta mata uangnya berubah")
        return 0

    if args.dari_arsip:
        tickers = ({t.strip().upper() for t in args.tickers.split(",") if t.strip()}
                   if args.tickers else None)
        return jalankan_arsip(
            tickers,
            None if args.semua_arsip else args.tahun,
            None if args.semua_arsip else args.periode,
        )

    if args.tickers:
        tickers = [t.strip().upper() for t in args.tickers.split(",") if t.strip()]
    elif args.semua:
        tickers = list(DEFAULT_TICKERS)
    else:
        ap.error("Pakai --tickers AAA,BBB atau --semua")
        return 1

    if args.batas:
        tickers = tickers[: args.batas]

    bucket = "tahunan" if args.periode == "audit" else "kuartal"
    tanggal = tanggal_akhir(args.tahun, args.periode)

    # curl_cffi.Session, bukan requests.Session -- lihat catatan IMPERSONATE.
    sesi = cffi.Session(impersonate=IMPERSONATE)
    try:
        sesi.get(PEMANASAN, headers={"User-Agent": UA}, timeout=30)
        print(f"Mengambil daftar laporan {args.periode.upper()} {args.tahun} ...")
        daftar = ambil_daftar(sesi, args.tahun, args.periode)
    except Exception as e:  # noqa: BLE001
        print(f"Gagal mengambil daftar laporan: {e}", file=sys.stderr)
        # Pesan ini dulu berbunyi "endpoint IDX hanya terbuka dari IP rumahan,
        # 403 dari datacenter" — dan itu menyesatkan sampai memakan waktu nyata
        # (18 Agu 2026): ia menuntun pembacanya menyalahkan ALAMAT, lalu menunggu
        # atau pindah mesin, padahal 403 hari itu datang dari bentuk permintaan dan
        # alamatnya tak pernah berubah. Pesan galat yang menyebut satu sebab dengan
        # yakin lebih buruk daripada tak ada pesan sama sekali: ia menutup arah
        # pemeriksaan lain. Sekarang menyebut uji yang MEMBEDAKAN keduanya.
        print(
            "403 di sini biasanya BENTUK permintaan, bukan alamat IP. Uji dulu: buka URL "
            "yang sama di peramban. Kalau peramban menjawab 200, yang ditolak sidik jari "
            "permintaannya \u2014 perbaiki header (lihat _HDR_PERAMBAN) atau pakai "
            "curl_cffi impersonate=chrome124. Kalau peramban IKUT 403, barulah curigai "
            "alamat/IP.",
            file=sys.stderr,
        )
        return 1
    print(f"  {len(daftar)} emiten punya laporan pada periode ini -> target {len(tickers)} tiker\n")

    ok = kosong = gagal = dilewati = 0
    for i, kode in enumerate(tickers, 1):
        print(f"  [{i:>3}/{len(tickers)}] {kode:<8}", end="", flush=True)

        if not args.paksa and sudah_ada(kode, tanggal, bucket):
            dilewati += 1
            print(" - dilewati (sudah ada)")
            continue

        entri = daftar.get(kode)
        if entri is None:
            kosong += 1
            print(" - kosong (tak ada laporan periode ini)")
            continue

        att = cari_xlsx(entri)
        if att is None:
            kosong += 1
            print(" - kosong (tak ada lampiran xlsx)")
            continue

        try:
            konten = ambil_xlsx(sesi, att["File_Path"], kode, args.tahun, args.periode)
            wb = load_workbook(io.BytesIO(konten), data_only=True, read_only=True)
            data_periode, currency, _kurs = ekstrak(wb)
            wb.close()
        except Ditolak:
            gagal += 1
            print(" - GAGAL (ditolak berturut-turut) -- berhenti")
            break
        except Exception as e:  # noqa: BLE001
            gagal += 1
            print(f" - gagal ({e})")
            time.sleep(random.uniform(*JEDA))
            continue

        terisi = sum(1 for v in data_periode.values() if v is not None)
        if terisi == 0:
            kosong += 1
            print(" - kosong (semua ruas null)")
        else:
            # Penulisan bisa gagal sesaat di Windows (kunci antivirus/pencadangan).
            # Dicoba ulang sekali; kalau tetap gagal, emiten ini saja yang hilang.
            try:
                simpan(kode, tanggal, bucket, data_periode, currency)
            except OSError as e:
                time.sleep(1.5)
                try:
                    simpan(kode, tanggal, bucket, data_periode, currency)
                except OSError as e2:
                    gagal += 1
                    print(f" - gagal menulis ({e2})")
                    time.sleep(random.uniform(*JEDA))
                    continue
            ok += 1
            print(f" - ok ({currency}, {terisi}/{len(ALL_KEYS)} ruas terisi)")

        time.sleep(random.uniform(*JEDA))

    print(f"\nSelesai: {ok} berhasil, {kosong} kosong, {gagal} gagal, "
          f"{dilewati} dilewati dari {len(tickers)} emiten")

    # Penambal WAJIB jalan sesudah pemanennya (CLAUDE.md, kasus
    # `lengkapi_fundamental.py`): panen menulis ulang nilai periode dari nol,
    # jadi tambalan skala sebelumnya ikut terhapus. `perbaiki_skala_keuangan.py`
    # sebelumnya tak dipanggil dari mana pun -- ia menambal sekali lalu tiap
    # panen berikutnya merusaknya lagi, tanpa satu pun galat. Dipanggil di sini,
    # bukan diserahkan ke ingatan orang yang menjalankan panen.
    if ok:
        import perbaiki_skala_keuangan
        print("\nMenambal skala (perbaiki_skala_keuangan) ...")
        perbaiki_skala_keuangan.jalankan(tulis=True)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
