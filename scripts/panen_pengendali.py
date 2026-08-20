"""Peras informasi pemegang saham pengendali dari arsip XBRL → data-idx/json/pengendali.json.

**Nol jaringan.** Bahannya berkas XLSX yang SUDAH ada di `_arsip-mentah/keuangan_idx/`
(±6.679 berkas, hasil panen `panen_keuangan_idx.py`) — aturan proyek "simpan mentah
hasil panen" persis untuk kebutuhan seperti ini: menambah ruas jadi tak berbiaya
jaringan sama sekali.

## Yang PERLU diketahui sebelum memakai keluarannya

Ruas `Informasi pemegang saham pengendali` di sheet `1000000` itu **kategori, bukan
nama**. Diperiksa 20 Agu 2026 atas 120 berkas acak (2025/tw3): 119 terisi, dan
nilainya cuma enam macam —

    National Corporation · Individual WNI · Foreign Corporation ·
    National and Foreign Corporation · Indonesian Government · No Controlling Shareholder

Jadi berkas ini menjawab "pengendalinya JENIS apa", **bukan** "siapa namanya".
Seluruh sheet lain sudah disapu untuk mencari daftar nama pemegang saham (BBCA &
ASII, 47 sheet): tidak ada. Yang muncul cuma pos neraca "piutang/utang pemegang
saham", bukan identitas. Nama pengendali harus dari sumber lain (KSEI, prospektus).

## Kenapa yang diambil hanya laporan TERBARU per emiten

Kepemilikan berubah. Menampilkan kategori tanpa tanggal laporannya membuat angka
lama terbaca sebagai posisi hari ini, jadi tanggalnya ikut disimpan dan wajib
tampil di layar. Arsip ditelusuri dari periode terbaru ke terlama dan emiten yang
sudah ketemu dilewati — jadi yang dibuka ±1.000 berkas, bukan 6.679.

Skrip ini **hanya membaca** arsip dan menulis satu berkas baru. Ia tidak menyentuh
`data-idx/json/keuangan_idx/` sama sekali, jadi tak ada risiko re-parse `--paksa`
yang menghapus catatan tahunan.

Pakai:
  C:/Python314/python.exe scripts/panen_pengendali.py
"""
from __future__ import annotations

import json
import sys
import warnings
from datetime import datetime, timedelta, timezone
from pathlib import Path

import openpyxl

warnings.filterwarnings("ignore", module="openpyxl")

AKAR = Path(__file__).resolve().parent.parent
ARSIP = AKAR / "_arsip-mentah" / "keuangan_idx"
KELUARAN = AKAR / "data-idx" / "json" / "pengendali.json"
WIB = timezone(timedelta(hours=7))

# Urutan periode dari TERBARU ke terlama. Ditulis eksplisit, bukan hasil sort
# nama folder: "audit" dan "tw1..tw3" tak bisa diurutkan bersama secara leksikal
# (`audit` < `tw1`), dan urutan yang salah di sini membuat seluruh berkas
# berjangkar pada laporan lama tanpa satu pun galat.
URUTAN = [
    ("2026", "tw2"), ("2026", "tw1"),
    ("2025", "tw3"), ("2025", "tw2"), ("2025", "tw1"),
    ("2024", "audit"), ("2023", "audit"), ("2022", "audit"),
    ("2021", "audit"), ("2020", "audit"), ("2019", "audit"),
]

# Label kolom pertama sheet 1000000 → kunci keluaran.
LABEL = {
    "Kode entitas": "kode",
    "Informasi pemegang saham pengendali": "jenis",
    "Tanggal akhir periode berjalan": "tanggal",
    "Periode penyampaian laporan keuangan": "periode",
}


def baca(path: Path) -> dict | None:
    """Ruas pengendali + tanggal dari sheet 1000000. None kalau sheetnya tak ada."""
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as e:  # noqa: BLE001 — berkas rusak/terpotong dilewati, bukan menghentikan panen
        print(f"  lewat {path.name}: {e}", file=sys.stderr)
        return None
    try:
        if "1000000" not in wb.sheetnames:
            return None
        hasil: dict[str, str] = {}
        for row in wb["1000000"].iter_rows(min_row=1, max_row=45, max_col=3, values_only=True):
            if not row or not row[0]:
                continue
            kunci = LABEL.get(str(row[0]).strip())
            if kunci and len(row) > 1 and row[1] is not None:
                nilai = str(row[1]).strip()
                if nilai:
                    hasil[kunci] = nilai
        return hasil or None
    finally:
        wb.close()


def main() -> int:
    if not ARSIP.is_dir():
        print(f"Arsip mentah tak ada: {ARSIP}", file=sys.stderr)
        return 1

    emiten: dict[str, dict] = {}
    dibuka = 0
    for tahun, periode in URUTAN:
        folder = ARSIP / tahun / periode
        if not folder.is_dir():
            continue
        baru = 0
        for berkas in sorted(folder.glob("*.xlsx")):
            kode_berkas = berkas.stem.upper()
            if kode_berkas in emiten:
                continue  # sudah dapat dari laporan yang lebih baru
            dibuka += 1
            isi = baca(berkas)
            if not isi or not isi.get("jenis"):
                continue
            kode = (isi.get("kode") or kode_berkas).upper()
            if kode in emiten:
                continue
            emiten[kode] = {
                "jenis": isi["jenis"],
                # Tanggal akhir periode laporan yang jadi sumbernya. Wajib
                # tampil di layar — tanpa ini kategori lama terbaca sebagai
                # posisi hari ini.
                "tanggal": isi.get("tanggal"),
                "periode": isi.get("periode"),
                "arsip": f"{tahun}/{periode}",
            }
            baru += 1
        print(f"{tahun}/{periode}: +{baru} emiten (total {len(emiten)})")

    if not emiten:
        print("Nol emiten terbaca — berkas lama TIDAK ditimpa.", file=sys.stderr)
        return 1

    isi = {
        "diperbarui": datetime.now(WIB).isoformat(timespec="seconds"),
        "sumber": "IDX GetFinancialReport (XBRL) sheet 1000000, arsip _arsip-mentah/keuangan_idx/",
        "catatan": (
            "Nilai `jenis` adalah KATEGORI pengendali menurut taksonomi XBRL bursa, "
            "bukan nama pemegang saham. Laporan resmi tidak memuat nama."
        ),
        "n": len(emiten),
        "emiten": dict(sorted(emiten.items())),
    }
    KELUARAN.parent.mkdir(parents=True, exist_ok=True)
    KELUARAN.write_text(json.dumps(isi, ensure_ascii=False, indent=1), encoding="utf-8")

    ragam: dict[str, int] = {}
    for v in emiten.values():
        ragam[v["jenis"]] = ragam.get(v["jenis"], 0) + 1
    print(f"\nOK -> {KELUARAN} ({len(emiten)} emiten, {dibuka} berkas dibuka)")
    for nama, n in sorted(ragam.items(), key=lambda x: -x[1]):
        print(f"  {n:4d}  {nama}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
